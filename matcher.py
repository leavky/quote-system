from __future__ import annotations

import math
import re
import sys
import unicodedata
from dataclasses import dataclass, asdict
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook


def base_dir() -> Path:
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent


BASE_DIR = base_dir()
UPLOAD_DIR = BASE_DIR / "uploads"
OUTPUT_DIR = BASE_DIR / "outputs"
DEFAULT_PRICE = BASE_DIR / "检测报价表.xlsx"
QUOTE_TEMPLATE = BASE_DIR / "报价清单导入模板.xlsx"


PRICE_COLUMNS = {
    "序号",
    "检测项目",
    "检测材料",
    "检测参数",
    "单位",
    "单价（元）",
    "单价",
    "备注",
    "报价编号",
    "检测项目别名",
    "检测参数别名",
}

QUOTE_TEMPLATE_COLUMNS = {
    "序号",
    "样品名称",
    "检测材料",
    "检测项目",
    "检测参数",
    "具体检测项目",
    "组/点数",
    "备注",
}

QUOTE_HINT_COLUMNS = {
    "序号",
    "样品名称",
    "检测材料",
    "检测项目",
    "检测参数",
    "具体检测项目",
    "组/点数",
    "备注",
}

STOP_WORDS = ("检测", "试验", "项目", "参数", "性能", "含量", "测定")

_PRICE_CACHE: dict[str, list["PriceItem"]] = {}


@dataclass
class PriceItem:
    id: str
    sheet: str
    row_number: int
    seq: str
    category: str
    material: str
    parameter: str
    unit: str
    price: float | None
    raw_price: str
    remark: str
    code: str
    project_aliases: list[str]
    parameter_aliases: list[str]
    aliases: list[str]
    search_text: str
    price_rule: dict[str, Any] | None


@dataclass
class QuoteLine:
    id: str
    sheet: str
    row_number: int
    seq: str
    sample_name: str
    project_name: str
    parameter: str
    unit: str
    quantity: float | None
    source_price: float | None
    source_total: float | None
    remark: str
    search_text: str


def quote_line_from_match(match: dict[str, Any]) -> QuoteLine:
    return QuoteLine(
        id=match.get("id", ""),
        sheet=match.get("sheet", ""),
        row_number=int(match.get("row_number") or 0),
        seq=match.get("seq", ""),
        sample_name=match.get("sample_name", ""),
        project_name=match.get("project_name", ""),
        parameter=match.get("parameter", ""),
        unit=match.get("unit", ""),
        quantity=match.get("quantity"),
        source_price=match.get("source_price"),
        source_total=match.get("source_total"),
        remark=match.get("remark", ""),
        search_text=match.get("search_text", ""),
    )


def normalize_text(value: Any, *, drop_stop_words: bool = False) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    text = unicodedata.normalize("NFKC", str(value)).strip().lower()
    text = text.replace("（", "(").replace("）", ")")
    text = re.sub(r"\s+", "", text)
    text = re.sub(r"[，,、;；/\\|+＋]", " ", text)
    text = re.sub(r"[()（）【】\[\]{}:：\-—_]", "", text)
    text = re.sub(r"\s+", " ", text).strip()
    if drop_stop_words:
        for word in STOP_WORDS:
            text = text.replace(word, "")
    return text


def compact_text(value: Any) -> str:
    return normalize_text(value, drop_stop_words=True).replace(" ", "")


def cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    text = str(value).strip()
    if text.endswith(".0") and text[:-2].isdigit():
        return text[:-2]
    return text


def to_number(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    text = cell_text(value)
    if not text:
        return None
    normalized = text.replace(",", "")
    if not re.fullmatch(r"-?\d+(?:\.\d+)?", normalized):
        return None
    return float(normalized)


def split_aliases(value: Any) -> list[str]:
    text = cell_text(value)
    if not text:
        return []
    parts = re.split(r"[,，、;；/|｜\n]+", text)
    return [p.strip() for p in parts if p.strip()]


def make_line_alias(sample_name: str, parameter: str) -> str:
    sample = cell_text(sample_name)
    param = cell_text(parameter)
    if sample and param:
        return f"{sample}:{param}"
    return param or sample


def parse_price_rule(raw_price: str, remark: str = "") -> dict[str, Any] | None:
    return None  # 暂不解析分档计价规则
    text = cell_text(raw_price)
    if not text:
        return None
    tiers = []
    for match in re.finditer(r"(\d+(?:\.\d+)?)\s*[\(（]([^()（）]+)[\)）]", text):
        tiers.append(
            {
                "label": match.group(2).strip(),
                "price": float(match.group(1)),
                "source": match.group(0),
            }
        )
    if not tiers:
        return None
    multipliers = []
    remark_text = cell_text(remark)
    if re.search(r"[Ⅲ三3]\s*级.*以上.*双倍", remark_text):
        multipliers.append(
            {
                "label": "Ⅲ级钢级以上双倍计费",
                "factor": 2,
                "trigger": "grade_gte_3",
                "source": remark_text,
            }
        )
    return {
        "type": "tiered",
        "raw_price": text,
        "tiers": tiers,
        "multipliers": multipliers,
    }


def parse_bar_diameter(text: str) -> float | None:
    normalized = unicodedata.normalize("NFKC", text)
    patterns = [
        r"[Φφ]\s*(\d+(?:\.\d+)?)",
        r"\bD\s*(\d+(?:\.\d+)?)\b",
        r"直径\s*(\d+(?:\.\d+)?)",
    ]
    for pattern in patterns:
        found = re.search(pattern, normalized, re.IGNORECASE)
        if found:
            return float(found.group(1))
    return None


def is_grade_gte_3(text: str) -> bool:
    normalized = unicodedata.normalize("NFKC", text).lower()
    return bool(re.search(r"(ⅲ|iii|三|3|三级|hrb400|hrb500|hrbf400|hrbf500)", normalized))


def tier_matches_condition(tier_label: str, context: str) -> bool:
    diameter = parse_bar_diameter(context)
    if diameter is None:
        return False
    label = unicodedata.normalize("NFKC", tier_label).replace(" ", "")
    if re.search(r"Φ?[<＜]\s*25", label):
        return diameter < 25
    if re.search(r"Φ?[≥>=]\s*25", label) or re.search(r"Φ?大于等于25", label):
        return diameter >= 25
    return False


def evaluate_price_rule(rule: dict[str, Any] | None, line: QuoteLine, item: PriceItem) -> dict[str, Any] | None:
    if not rule:
        return None
    context = " ".join([line.sample_name, line.project_name, line.parameter, line.remark])
    selected_tier = None
    for tier in rule.get("tiers", []):
        if tier_matches_condition(tier.get("label", ""), context):
            selected_tier = tier
            break
    factor = 1.0
    applied_multipliers = []
    for multiplier in rule.get("multipliers", []):
        if multiplier.get("trigger") == "grade_gte_3" and is_grade_gte_3(context):
            factor *= float(multiplier.get("factor", 1))
            applied_multipliers.append(multiplier)
    if not selected_tier:
        return {
            "requires_selection": True,
            "options": rule.get("tiers", []),
            "multipliers": rule.get("multipliers", []),
            "message": "复合报价，需选择规格后计算",
        }
    unit_price = round(float(selected_tier["price"]) * factor, 2)
    qty = line.quantity or 1
    explanation = selected_tier["label"]
    if applied_multipliers:
        explanation += "；" + "；".join(m["label"] for m in applied_multipliers)
    return {
        "requires_selection": False,
        "selected": selected_tier,
        "base_price": selected_tier["price"],
        "factor": factor,
        "unit_price": unit_price,
        "total": round(unit_price * qty, 2),
        "explanation": explanation,
        "options": rule.get("tiers", []),
        "multipliers": rule.get("multipliers", []),
    }


def similarity(left: str, right: str) -> float:
    left_c = compact_text(left)
    right_c = compact_text(right)
    if not left_c or not right_c:
        return 0.0
    ratio = SequenceMatcher(None, left_c, right_c).ratio()
    if left_c in right_c or right_c in left_c:
        ratio = max(ratio, min(len(left_c), len(right_c)) / max(len(left_c), len(right_c)) + 0.18)
    return min(ratio * 100, 100)


def find_header_row(df: pd.DataFrame, hints: set[str]) -> int | None:
    best_idx = None
    best_score = 0
    for idx in range(min(len(df), 15)):
        values = {cell_text(v) for v in df.iloc[idx].tolist()}
        score = len(values & hints)
        if score > best_score:
            best_score = score
            best_idx = idx
    return best_idx if best_score >= 2 else None


def dataframe_from_sheet(path: Path, sheet_name: str, hints: set[str]) -> tuple[pd.DataFrame, int]:
    raw = pd.read_excel(path, sheet_name=sheet_name, header=None)
    header_idx = find_header_row(raw, hints)
    if header_idx is None:
        raise ValueError(f"{path.name} / {sheet_name} 未找到可识别表头")
    headers = [cell_text(v) or f"未命名{idx + 1}" for idx, v in enumerate(raw.iloc[header_idx].tolist())]
    df = raw.iloc[header_idx + 1 :].copy()
    df.columns = headers
    df = df.dropna(how="all")
    return df, header_idx + 2


def read_price_book(path: str | Path) -> list[PriceItem]:
    path = Path(path)
    workbook = pd.ExcelFile(path)
    items: list[PriceItem] = []
    for sheet in workbook.sheet_names:
        df, first_excel_row = dataframe_from_sheet(path, sheet, PRICE_COLUMNS)
        for col in ("检测项目", "检测材料"):
            if col in df.columns:
                df[col] = df[col].ffill()
        for offset, row in df.iterrows():
            parameter = cell_text(row.get("检测参数"))
            category = cell_text(row.get("检测项目"))
            material = cell_text(row.get("检测材料"))
            price = to_number(row.get("单价（元）", row.get("单价")))
            raw_price = cell_text(row.get("单价（元）", row.get("单价")))
            remark = cell_text(row.get("备注"))
            code = cell_text(row.get("报价编号"))
            if not parameter or not raw_price:
                continue
            row_number = int(offset) + 1
            project_aliases = split_aliases(row.get("检测项目别名"))
            parameter_aliases = split_aliases(row.get("检测参数别名"))
            aliases = [*project_aliases, *parameter_aliases]
            searchable = " ".join([sheet, category, material, parameter, *aliases])
            price_rule = parse_price_rule(raw_price, remark)
            items.append(
                PriceItem(
                    id=code or f"{sheet}-{row_number}",
                    sheet=sheet,
                    row_number=row_number,
                    seq=cell_text(row.get("序号")),
                    category=category,
                    material=material,
                    parameter=parameter,
                    unit=cell_text(row.get("单位")),
                    price=price,
                    raw_price=raw_price,
                    remark=remark,
                    code=code,
                    project_aliases=project_aliases,
                    parameter_aliases=parameter_aliases,
                    aliases=aliases,
                    search_text=searchable,
                    price_rule=price_rule,
                )
            )
    return items


def load_price_items(path: str | Path) -> list[PriceItem]:
    key = str(Path(path).resolve())
    if key not in _PRICE_CACHE:
        _PRICE_CACHE[key] = read_price_book(path)
    return _PRICE_CACHE[key]


def clear_price_cache(path: str | Path) -> None:
    _PRICE_CACHE.pop(str(Path(path).resolve()), None)


def find_header_cells_in_sheet(sheet, hints: set[str]) -> tuple[int | None, dict[str, int]]:
    best_row = None
    best_score = 0
    best_headers: dict[str, int] = {}
    max_row = min(sheet.max_row, 15)
    for row_idx in range(1, max_row + 1):
        headers = {}
        values = set()
        for cell in sheet[row_idx]:
            text = cell_text(cell.value)
            if text:
                headers[text] = cell.column
                values.add(text)
        score = len(values & hints)
        if score > best_score:
            best_score = score
            best_row = row_idx
            best_headers = headers
    if best_score < 2:
        return None, {}
    return best_row, best_headers


def append_price_aliases(
    path: str | Path,
    item_data: dict[str, Any],
    project_alias: str = "",
    parameter_alias: str = "",
) -> dict[str, Any]:
    path = Path(path)
    sheet_name = item_data.get("sheet")
    row_number = int(item_data.get("row_number") or 0)
    if not sheet_name or row_number <= 0:
        return {"updated": False, "message": "缺少报价库行定位信息"}

    workbook = load_workbook(path)
    if sheet_name not in workbook.sheetnames:
        return {"updated": False, "message": f"报价库中未找到分类：{sheet_name}"}
    sheet = workbook[sheet_name]
    header_row, headers = find_header_cells_in_sheet(sheet, PRICE_COLUMNS)
    if header_row is None:
        return {"updated": False, "message": "未找到报价库表头"}

    row_number = resolve_price_item_row(sheet, headers, header_row, row_number, item_data)
    if row_number is None:
        return {"updated": False, "message": "未能定位报价库对应行"}

    project_result = append_alias_to_column(
        sheet,
        header_row,
        headers,
        row_number,
        "检测项目别名",
        project_alias,
        [item_data.get("category"), item_data.get("material")],
    )
    parameter_result = append_alias_to_column(
        sheet,
        header_row,
        headers,
        row_number,
        "检测参数别名",
        parameter_alias,
        [item_data.get("parameter")],
    )
    updated = project_result["updated"] or parameter_result["updated"]
    if updated:
        workbook.save(path)
        clear_price_cache(path)
    messages = [result["message"] for result in (project_result, parameter_result) if result["message"]]
    return {
        "updated": updated,
        "message": "；".join(messages) or "没有可写入的别名",
        "project_aliases": project_result["aliases"],
        "parameter_aliases": parameter_result["aliases"],
    }


def append_alias_to_column(
    sheet,
    header_row: int,
    headers: dict[str, int],
    row_number: int,
    column_name: str,
    alias: str,
    standard_values: list[Any],
) -> dict[str, Any]:
    alias = cell_text(alias)
    if not alias:
        return {"updated": False, "message": "", "aliases": []}
    alias_col = headers.get(column_name)
    if alias_col is None:
        alias_col = sheet.max_column + 1
        sheet.cell(row=header_row, column=alias_col, value=column_name)
        headers[column_name] = alias_col

    cell = sheet.cell(row=row_number, column=alias_col)
    existing_aliases = split_aliases(cell.value)
    existing_norms = {compact_text(value) for value in existing_aliases}
    alias_norm = compact_text(alias)
    standard_norms = {compact_text(value) for value in standard_values if cell_text(value)}
    if not alias_norm or alias_norm in existing_norms or alias_norm in standard_norms:
        return {"updated": False, "message": f"{column_name}已存在或无需写入", "aliases": existing_aliases}

    updated_aliases = [*existing_aliases, alias]
    cell.value = "/".join(updated_aliases)
    return {"updated": True, "message": f"{column_name}已写入", "aliases": updated_aliases}


def resolve_price_item_row(sheet, headers: dict[str, int], header_row: int, row_number: int, item_data: dict[str, Any]) -> int | None:
    if row_number_matches_price_item(sheet, headers, row_number, item_data):
        return row_number

    last_category = ""
    last_material = ""
    for row_idx in range(header_row + 1, sheet.max_row + 1):
        category = cell_text(sheet.cell(row=row_idx, column=headers.get("检测项目", 0)).value) if headers.get("检测项目") else ""
        material = cell_text(sheet.cell(row=row_idx, column=headers.get("检测材料", 0)).value) if headers.get("检测材料") else ""
        if category:
            last_category = category
        if material:
            last_material = material
        if row_number_matches_price_item(sheet, headers, row_idx, item_data, category=last_category, material=last_material):
            return row_idx
    return None


def row_number_matches_price_item(
    sheet,
    headers: dict[str, int],
    row_number: int,
    item_data: dict[str, Any],
    *,
    category: str | None = None,
    material: str | None = None,
) -> bool:
    if row_number < 1 or row_number > sheet.max_row:
        return False
    parameter_col = headers.get("检测参数")
    if not parameter_col:
        return False
    parameter = cell_text(sheet.cell(row=row_number, column=parameter_col).value)
    if compact_text(parameter) != compact_text(item_data.get("parameter")):
        return False

    seq_col = headers.get("序号")
    if seq_col and cell_text(item_data.get("seq")):
        seq = cell_text(sheet.cell(row=row_number, column=seq_col).value)
        if seq and compact_text(seq) != compact_text(item_data.get("seq")):
            return False

    price_col = headers.get("单价（元）") or headers.get("单价")
    if price_col and cell_text(item_data.get("raw_price")):
        raw_price = cell_text(sheet.cell(row=row_number, column=price_col).value)
        if raw_price and compact_text(raw_price) != compact_text(item_data.get("raw_price")):
            return False

    if category is not None and cell_text(item_data.get("category")):
        if compact_text(category) != compact_text(item_data.get("category")):
            return False
    if material is not None and cell_text(item_data.get("material")):
        if compact_text(material) != compact_text(item_data.get("material")):
            return False
    return True


def search_blob(item: PriceItem) -> str:
    return " ".join([item.code, item.category, item.material, item.parameter, item.remark, " ".join(item.aliases)])


def project_blob(item: PriceItem) -> str:
    return " ".join([item.category, item.material, *item.project_aliases])


def parameter_blob(item: PriceItem) -> str:
    return " ".join([item.parameter, *item.parameter_aliases])


def line_material_text(line: QuoteLine) -> str:
    return line.sample_name


def line_detection_text(line: QuoteLine) -> str:
    return line.parameter or line.project_name


def score_line_against_item(line: QuoteLine, item: PriceItem) -> tuple[float, str]:
    line_material = line_material_text(line)
    line_detection = line_detection_text(line)
    material_score = max(similarity(line_material, item.material), similarity(line_material, item.category))
    project_score = similarity(line.project_name, project_blob(item)) if line.project_name else 0.0
    parameter_score = similarity(line_detection, parameter_blob(item))
    context_score = max(material_score, project_score)

    material_exact = compact_text(line_material) and compact_text(line_material) in compact_text(project_blob(item))
    parameter_exact = compact_text(line_detection) and compact_text(line_detection) in compact_text(parameter_blob(item))
    if parameter_exact and (material_exact or context_score >= 70):
        return 98.0, "材料+参数匹配"
    if parameter_exact:
        return min(62 + context_score * 0.28, 88), "参数匹配"
    if material_exact:
        return min(66 + parameter_score * 0.28 + project_score * 0.06, 92), "材料匹配"

    score = parameter_score * 0.5 + material_score * 0.38 + project_score * 0.12
    return score, "三列模糊匹配"


def search_price_items(path: str | Path, query: str, limit: int | None = 50, mode: str = "fuzzy") -> list[dict[str, Any]]:
    items = load_price_items(path)
    q = compact_text(query)
    ranked = []
    for item in items:
        blob = search_blob(item)
        score = 0.0
        if not q:
            score = 1.0
        else:
            text = compact_text(blob)
            if q in text:
                score = 90 + min(len(q) / max(len(text), 1) * 10, 10)
            elif mode != "exact":
                score = similarity(query, blob)
            else:
                continue
        ranked.append((score, item))
    ranked.sort(key=lambda x: x[0], reverse=True)
    results = []
    if limit is not None:
        ranked = ranked[:limit]
    for score, item in ranked:
        if q and score < 20:
            continue
        data = asdict(item)
        data["score"] = round(score, 1)
        results.append(data)
    return results


def score_price_item(line: QuoteLine, item: PriceItem, query: str = "", mode: str = "fuzzy") -> tuple[float, str] | None:
    score, method = exact_or_alias_score(line, item)
    if not score:
        score, method = score_line_against_item(line, item)
    if query:
        blob = search_blob(item)
        query_c = compact_text(query)
        if mode == "exact" and query_c not in compact_text(blob):
            return None
        query_score = 100.0 if query_c in compact_text(blob) else similarity(query, blob)
        score = max(score, query_score * 0.9)
    return score, method


def rank_price_items_for_line(
    price_path: str | Path,
    line: QuoteLine,
    query: str = "",
    tab: str = "推荐",
    limit: int | None = None,
    mode: str = "fuzzy",
) -> list[dict[str, Any]]:
    items = load_price_items(price_path)
    ranked = []
    query_text = cell_text(query)
    for item in items:
        if tab and tab != "推荐" and item.sheet != tab:
            continue
        scored = score_price_item(line, item, query_text, mode=mode)
        if scored is None:
            continue
        score, method = scored
        if query_text and score < 20:
            continue
        ranked.append((score, method, item))
    if tab and tab != "推荐":
        ranked.sort(key=lambda x: (x[2].row_number, x[2].seq))
    else:
        ranked.sort(key=lambda x: x[0], reverse=True)
    if limit is not None:
        ranked = ranked[:limit]
    return [
        {
            **asdict(item),
            "score": round(score, 1),
            "method": method,
        }
        for score, method, item in ranked
    ]


def read_quote_book(path: str | Path) -> list[QuoteLine]:
    path = Path(path)
    if path.suffix.lower() == ".xls":
        raise ValueError("暂不支持 .xls，请先另存为 .xlsx 后再上传")
    workbook = pd.ExcelFile(path)
    lines: list[QuoteLine] = []
    for sheet in workbook.sheet_names:
        df, first_excel_row = dataframe_from_sheet(path, sheet, QUOTE_HINT_COLUMNS)
        for col in ("样品名称", "检测材料", "检测项名称", "单位工程", "组/点数", "数量"):
            if col in df.columns:
                df[col] = df[col].ffill()
        for offset, row in df.iterrows():
            sample = cell_text(row.get("检测材料")) or cell_text(row.get("样品名称")) or cell_text(row.get("检测项名称"))
            project = cell_text(row.get("检测项目"))
            parameter = cell_text(row.get("检测参数")) or cell_text(row.get("具体检测项目")) or project
            if not sample and not parameter:
                continue
            source_price = to_number(row.get("单价（元）", row.get("单价")))
            quantity = to_number(row.get("数量", row.get("组/点数")))
            source_total = to_number(row.get("合价（元）", row.get("合价")))
            row_number = int(offset) + first_excel_row
            search_text = " ".join([sample, project, parameter, cell_text(row.get("备注"))])
            lines.append(
                QuoteLine(
                    id=f"{sheet}-{row_number}",
                    sheet=sheet,
                    row_number=row_number,
                    seq=str(len(lines) + 1),
                    sample_name=sample,
                    project_name=project,
                    parameter=parameter,
                    unit=cell_text(row.get("单位")),
                    quantity=quantity,
                    source_price=source_price,
                    source_total=source_total,
                    remark=cell_text(row.get("备注")),
                    search_text=search_text,
                )
            )
    return lines


def exact_or_alias_score(line: QuoteLine, item: PriceItem) -> tuple[float, str]:
    line_param = compact_text(line.parameter or line.project_name)
    line_all = compact_text(line.search_text)
    line_pair = compact_text(make_line_alias(line_material_text(line), line_detection_text(line)))
    item_param = compact_text(item.parameter)
    line_sample = compact_text(line_material_text(line))
    material_score = max(similarity(line_material_text(line), item.material), similarity(line_material_text(line), item.category))
    project_alias_match = any(
        alias_c and (line_sample == alias_c or alias_c in line_all)
        for alias_c in (compact_text(alias) for alias in item.project_aliases)
    )
    if line_param and line_param == item_param:
        if material_score >= 70 or project_alias_match:
            return min(88 + material_score * 0.12, 100), "材料+参数匹配"
        return min(62 + material_score * 0.28, 88), "参数匹配"
    for alias in item.parameter_aliases:
        alias_c = compact_text(alias)
        if alias_c and (line_pair == alias_c or line_param == alias_c or alias_c in line_all):
            base = 90 if project_alias_match or material_score >= 60 else 62
            return min(base + material_score * 0.1, 100), "参数别名匹配"
    for alias in item.aliases:
        alias_c = compact_text(alias)
        if alias_c and (line_pair == alias_c or line_param == alias_c or alias_c in line_all):
            base = 86 if project_alias_match or material_score >= 60 else 62
            return min(base + material_score * 0.1, 100), "别名匹配"
    return 0, ""


def match_line(line: QuoteLine, price_items: list[PriceItem]) -> dict[str, Any]:
    best: tuple[float, str, PriceItem | None] = (0, "", None)
    candidates = []
    for item in price_items:
        score, method = exact_or_alias_score(line, item)
        if not score:
            score, method = score_line_against_item(line, item)
        candidates.append((score, method, item))
        if score > best[0]:
            best = (score, method, item)
    candidates.sort(key=lambda x: x[0], reverse=True)
    score, method, item = best
    status = "自动匹配" if score >= 90 else "待确认" if score >= 65 else "未匹配"
    result = {
        **asdict(line),
        "match_status": status,
        "match_score": round(score, 1),
        "match_method": method if item else "",
        "matched": asdict(item) if item else None,
        "top_candidates": [
            {
                "score": round(candidate_score, 1),
                "method": candidate_method,
                "item": asdict(candidate_item),
            }
            for candidate_score, candidate_method, candidate_item in candidates[:5]
        ],
    }
    if item:
        qty = line.quantity or 1
        rule_result = evaluate_price_rule(item.price_rule, line, item)
        result["price_rule_result"] = rule_result
        result["matched_price"] = rule_result.get("unit_price") if rule_result and not rule_result.get("requires_selection") else item.price
        result["matched_price_text"] = item.raw_price
        result["matched_code"] = item.code or item.parameter or item.id
        result["matched_label"] = " / ".join(x for x in [item.category, item.material or item.parameter] if x)
        if rule_result and not rule_result.get("requires_selection"):
            result["calculated_total"] = rule_result.get("total")
            result["price_explanation"] = rule_result.get("explanation")
        elif rule_result and rule_result.get("requires_selection"):
            result["calculated_total"] = None
            result["price_explanation"] = rule_result.get("message")
            if status == "自动匹配":
                result["match_status"] = "待确认"
        else:
            result["calculated_total"] = round(qty * item.price, 2) if item.price is not None else None
            result["price_explanation"] = ""
    else:
        result["matched_price"] = None
        result["matched_price_text"] = None
        result["matched_code"] = None
        result["matched_label"] = None
        result["calculated_total"] = None
        result["price_rule_result"] = None
        result["price_explanation"] = ""
    return result


def match_quote(price_path: str | Path, quote_path: str | Path) -> dict[str, Any]:
    price_items = read_price_book(price_path)
    quote_lines = read_quote_book(quote_path)
    matches = [match_line(line, price_items) for line in quote_lines]
    summary = {
        "price_items": len(price_items),
        "quote_lines": len(quote_lines),
        "auto": sum(1 for m in matches if m["match_status"] == "自动匹配"),
        "review": sum(1 for m in matches if m["match_status"] == "待确认"),
        "unmatched": sum(1 for m in matches if m["match_status"] == "未匹配"),
    }
    return {"summary": summary, "matches": matches}


def export_matches(matches: list[dict[str, Any]], output_path: str | Path) -> Path:
    rows = []
    for seq, m in enumerate(matches, 1):
        item = m.get("matched") or {}
        suggested_price = m.get("matched_price_text") or (
            item.get("raw_price") if item else None
        )
        qty = m.get("quantity")
        calc_total = m.get("calculated_total")
        rows.append(
            {
                "序号": m.get("seq") or seq,
                "样品名称": m.get("sample_name"),
                "检测项目": m.get("parameter") or m.get("project_name"),
                "组/点数": qty,
                "单价": suggested_price,
                "合价": calc_total,
                "备注": item.get("remark") or m.get("matched_remark"),
                "报价编号": m.get("matched_code") or item.get("code"),
            }
        )
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    df = pd.DataFrame(rows)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="匹配结果")
        sheet = writer.book["匹配结果"]
        sheet.freeze_panes = "A2"
        for column_cells in sheet.columns:
            max_len = max(len(cell_text(cell.value)) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_len + 2, 10), 32)
    return output_path


def create_quote_template(output_path: str | Path) -> Path:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    columns = ["序号", "检测材料", "检测参数", "组/点数", "备注"]
    df = pd.DataFrame(columns=columns)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="报价清单")
        sheet = writer.book["报价清单"]
        sheet.freeze_panes = "A2"
        widths = {"A": 10, "B": 20, "C": 28, "D": 12, "E": 24}
        for col, width in widths.items():
            sheet.column_dimensions[col].width = width
    return output_path
