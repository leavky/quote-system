from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from settlement import build_match_result, export_result, reprice_result


BASE_DIR = Path(__file__).resolve().parents[1]
LEDGER = BASE_DIR / "zc检测费用台帐明细20260725.xls"
PRICE = BASE_DIR / "检测项目结算计费价格汇总.xlsx"


class SettlementMatcherTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.result = build_match_result(LEDGER, PRICE)

    def test_price_systems_are_f_to_j(self):
        self.assertEqual(
            [system["name"] for system in self.result["systems"]],
            ["25年协会价", "14年信息价", "宁大内部价", "其他价格", "预留备用"],
        )

    def test_exact_pair_uses_first_available_price(self):
        record = next(
            row
            for row in self.result["records"]
            if row["report_category"] == "混凝土抗压" and row["billing_item"] == "试块抗压（标养）"
        )
        self.assertEqual(record["status"], "已匹配")
        self.assertEqual(record["matched_code"], "HNT002")
        self.assertEqual(record["selected_system"], "25年协会价")
        self.assertEqual(record["selected_price"], 101.0)

    def test_priority_can_be_reordered(self):
        result = build_match_result(LEDGER, PRICE)
        priority = [
            {"name": "其他价格", "enabled": True},
            {"name": "25年协会价", "enabled": True},
            {"name": "14年信息价", "enabled": True},
            {"name": "宁大内部价", "enabled": True},
            {"name": "预留备用", "enabled": False},
        ]
        reprice_result(result, priority)
        record = next(
            row
            for row in result["records"]
            if row["report_category"] == "混凝土抗压" and row["billing_item"] == "试块抗压（标养）"
        )
        self.assertEqual(record["selected_system"], "其他价格")
        self.assertEqual(record["selected_price"], 101.3)

    def test_export_contains_formula_driven_amounts(self):
        with tempfile.TemporaryDirectory() as temporary:
            output = Path(temporary) / "result.xlsx"
            export_result(self.result, output, ledger_name=LEDGER.name, price_name=PRICE.name)
            workbook = load_workbook(output, data_only=False)
            self.assertEqual(workbook.sheetnames, ["结算汇总", "结算明细"])
            detail = workbook["结算明细"]
            headers = {detail.cell(2, column).value: column for column in range(1, detail.max_column + 1)}
            amount_column = headers["结算金额(元)"]
            formula_cells = [detail.cell(row, amount_column).value for row in range(3, detail.max_row + 1)]
            self.assertTrue(any(isinstance(value, str) and value.startswith("=IF(") for value in formula_cells))

    def test_empty_filter_exports_no_detail_rows(self):
        with tempfile.TemporaryDirectory() as temporary:
            output = Path(temporary) / "empty.xlsx"
            export_result(self.result, output, row_ids=[])
            workbook = load_workbook(output, data_only=False)
            self.assertEqual(workbook["结算明细"].max_row, 2)


if __name__ == "__main__":
    unittest.main()
