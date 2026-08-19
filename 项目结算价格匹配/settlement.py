from __future__ import annotations

import math
import re
import unicodedata
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


LEDGER_HINTS = {"委托日期", "报告编号", "报告类别", "工程名称", "计费项目", "数量"}
PRICE_HINTS = {"序号", "分类", "报告类别", "计费项目", "计费项目编号"}
RESULT_HEADERS = [
    "匹配状态",
    "计费项目编号",
    "价格表行",
    "匹配价格体系",
    "结算单价(元)",
    "结算数量",
    "结算金额(元)",
]


@dataclass
class SheetData:
    name: str
    rows: list[list[Any]]


def cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def exact_key(value: Any) -> str:
    text = unicodedata.normalize("NFKC", cell_text(value))
    return re.sub(r"\s+", "", text).strip()


def number_value(value: Any) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        number = float(value)
        return None if math.isnan(number) else number
    text = unicodedata.normalize("NFKC", cell_text(value)).replace(",", "")
    if not re.fullmatch(r"-?\d+(?:\.\d+)?", text):
        return None
    return float(text)


def json_value(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.isoformat(timespec="seconds")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and math.isnan(value):
        return None
    return value


def unique_headers(values: list[Any]) -> list[str]:
    counts: dict[str, int] = defaultdict(int)
    headers = []
    for index, value in enumerate(values, 1):
        base = cell_text(value) or f"未命名{index}"
        counts[base] += 1
        headers.append(base if counts[base] == 1 else f"{base}_{counts[base]}")
    return headers


def base_header(header: str) -> str:
    return re.sub(r"_\d+$", "", header)


def find_header_row(rows: list[list[Any]], hints: set[str]) -> int | None:
    best_index = None
    best_score = 0
    for index, row in enumerate(rows[:20]):
        values = {cell_text(value) for value in row if cell_text(value)}
        score = len(values & hints)
        if score > best_score:
            best_score = score
            best_index = index
    return best_index if best_score >= 2 else None


def read_xlsx(path: Path) -> list[SheetData]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheets = []
    try:
        for sheet in workbook.worksheets:
            rows = [list(row) for row in sheet.iter_rows(values_only=True)]
            sheets.append(SheetData(sheet.title, rows))
    finally:
        workbook.close()
    return sheets


def read_xls(path: Path) -> list[SheetData]:
    try:
        import xlrd
    except ImportError as exc:
        raise RuntimeError("读取 .xls 需要 xlrd，请先执行 uv sync") from exc

    workbook = xlrd.open_workbook(path)
    sheets = []
    for source_sheet in workbook.sheets():
        rows: list[list[Any]] = []
        for row_index in range(source_sheet.nrows):
            row = []
            for cell in source_sheet.row(row_index):
                value = cell.value
                if cell.ctype == xlrd.XL_CELL_DATE:
                    value = xlrd.xldate_as_datetime(value, workbook.datemode)
                elif cell.ctype == xlrd.XL_CELL_BOOLEAN:
                    value = bool(value)
                elif cell.ctype in {xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK}:
                    value = None
                row.append(value)
            rows.append(row)
        sheets.append(SheetData(source_sheet.name, rows))
    return sheets


def read_sheets(path: str | Path) -> list[SheetData]:
    source = Path(path)
    suffix = source.suffix.lower()
    if suffix == ".xlsx":
        return read_xlsx(source)
    if suffix == ".xls":
        return read_xls(source)
    raise ValueError("只支持 .xls 或 .xlsx 文件")


def as_excel_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)) and 1 <= float(value) <= 100000:
        return (datetime(1899, 12, 30) + timedelta(days=float(value))).date()
    text = cell_text(value)
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y年%m月%d日"):
        try:
            return datetime.strptime(text[:10], fmt).date()
        except ValueError:
            pass
    return None


def read_ledger(path: str | Path) -> dict[str, Any]:
    records: list[dict[str, Any]] = []
    all_headers: list[str] = []
    recognized_sheets = 0

    for sheet in read_sheets(path):
        header_index = find_header_row(sheet.rows, LEDGER_HINTS)
        if header_index is None:
            continue
        recognized_sheets += 1
        headers = unique_headers(sheet.rows[header_index])
        for header in headers:
            if header not in all_headers:
                all_headers.append(header)

        for row_index, values in enumerate(sheet.rows[header_index + 1 :], header_index + 2):
            padded = list(values[: len(headers)]) + [None] * max(0, len(headers) - len(values))
            original = dict(zip(headers, padded))
            if not any(cell_text(value) for value in original.values()):
                continue
            report_category = original.get("报告类别")
            billing_item = original.get("计费项目")
            report_number = original.get("报告编号") or original.get("受理编号")
            if not any(cell_text(value) for value in (report_category, billing_item, report_number)):
                continue

            entrusted_date = as_excel_date(original.get("委托日期"))
            quantity = number_value(original.get("数量"))
            source_unit_price = number_value(original.get("单价(元)") or original.get("单价（元）"))
            records.append(
                {
                    "id": f"{sheet.name}-{row_index}",
                    "sheet": sheet.name,
                    "row_number": row_index,
                    "original": original,
                    "date_value": entrusted_date,
                    "date": entrusted_date.isoformat() if entrusted_date else "",
                    "report_number": cell_text(report_number),
                    "project_name": cell_text(original.get("工程名称")),
                    "unit_project": cell_text(original.get("单体工程")),
                    "client_name": cell_text(original.get("委托单位")),
                    "report_category": cell_text(report_category),
                    "billing_item": cell_text(billing_item),
                    "quantity": quantity,
                    "source_unit_price": source_unit_price,
                }
            )

    if not recognized_sheets:
        raise ValueError("台账中未找到包含“报告类别”和“计费项目”的表头")
    return {"headers": all_headers, "records": records}


def read_price_book(path: str | Path) -> dict[str, Any]:
    items: list[dict[str, Any]] = []
    systems: list[dict[str, Any]] = []
    recognized_sheets = 0

    for sheet in read_sheets(path):
        header_index = find_header_row(sheet.rows, PRICE_HINTS)
        if header_index is None:
            continue
        recognized_sheets += 1
        raw_headers = list(sheet.rows[header_index])
        headers = unique_headers(raw_headers)
        if len(headers) < 6:
            raise ValueError(f"价格表 {sheet.name} 的价格列不足，预期 F-J 列")

        sheet_systems = []
        for column_index in range(5, min(10, len(headers))):
            name = cell_text(raw_headers[column_index]) or f"价格体系{column_index + 1}"
            if name not in [system["name"] for system in systems]:
                systems.append({"name": name, "column_index": column_index, "non_empty": 0})
            sheet_systems.append((name, column_index))

        for row_index, values in enumerate(sheet.rows[header_index + 1 :], header_index + 2):
            padded = list(values[: len(headers)]) + [None] * max(0, len(headers) - len(values))
            row = dict(zip(headers, padded))
            report_category = cell_text(row.get("报告类别"))
            billing_item = cell_text(row.get("计费项目"))
            if not report_category and not billing_item:
                continue
            prices: dict[str, float | None] = {}
            raw_prices: dict[str, Any] = {}
            for name, column_index in sheet_systems:
                raw = padded[column_index] if column_index < len(padded) else None
                prices[name] = number_value(raw)
                raw_prices[name] = json_value(raw)
                if prices[name] is not None:
                    next(system for system in systems if system["name"] == name)["non_empty"] += 1
            items.append(
                {
                    "sheet": sheet.name,
                    "row_number": row_index,
                    "report_category": report_category,
                    "billing_item": billing_item,
                    "code": cell_text(row.get("计费项目编号")),
                    "category": cell_text(row.get("分类")),
                    "prices": prices,
                    "raw_prices": raw_prices,
                }
            )

    if not recognized_sheets:
        raise ValueError("价格汇总表中未找到包含“报告类别”和“计费项目”的表头")
    if not systems:
        raise ValueError("价格汇总表未识别到 F-J 价格体系")
    return {"items": items, "systems": systems}


def normalize_priority(systems: list[dict[str, Any]], requested: list[dict[str, Any]] | None = None) -> list[dict[str, Any]]:
    known = {system["name"] for system in systems}
    result: list[dict[str, Any]] = []
    seen: set[str] = set()
    for entry in requested or []:
        name = cell_text(entry.get("name"))
        if name in known and name not in seen:
            result.append({"name": name, "enabled": bool(entry.get("enabled", True))})
            seen.add(name)
    for system in systems:
        if system["name"] not in seen:
            result.append({"name": system["name"], "enabled": True})
    if not any(entry["enabled"] for entry in result):
        raise ValueError("至少启用一个价格体系")
    return result


def match_records(
    ledger_records: list[dict[str, Any]],
    price_items: list[dict[str, Any]],
    priority: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    price_index: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for item in price_items:
        price_index[(exact_key(item["report_category"]), exact_key(item["billing_item"]))].append(item)

    matched_records = []
    enabled_priority = [entry["name"] for entry in priority if entry["enabled"]]
    for source in ledger_records:
        record = dict(source)
        report_key = exact_key(record["report_category"])
        item_key = exact_key(record["billing_item"])
        candidates = price_index.get((report_key, item_key), []) if report_key and item_key else []
        selected_item = None
        selected_system = ""
        selected_price = None

        if not report_key or not item_key:
            status = "字段缺失"
        elif not candidates:
            status = "未匹配"
        elif len(candidates) > 1:
            status = "价格表重复"
        else:
            selected_item = candidates[0]
            for system_name in enabled_priority:
                value = selected_item["prices"].get(system_name)
                if value is not None:
                    selected_system = system_name
                    selected_price = value
                    break
            status = "已匹配" if selected_price is not None else "无可用价格"

        quantity = record.get("quantity")
        amount = round(selected_price * quantity, 2) if selected_price is not None and quantity is not None else None
        record.update(
            {
                "status": status,
                "matched_code": selected_item["code"] if selected_item else "",
                "price_sheet": selected_item["sheet"] if selected_item else "",
                "price_row_number": selected_item["row_number"] if selected_item else None,
                "selected_system": selected_system,
                "selected_price": selected_price,
                "settlement_amount": amount,
                "available_prices": selected_item["prices"] if selected_item else {},
                "candidate_count": len(candidates),
            }
        )
        matched_records.append(record)
    return matched_records


def summarize(records: list[dict[str, Any]]) -> dict[str, Any]:
    status_counts: dict[str, int] = defaultdict(int)
    total_amount = 0.0
    for record in records:
        status_counts[record["status"]] += 1
        if record.get("settlement_amount") is not None:
            total_amount += float(record["settlement_amount"])
    return {
        "total": len(records),
        "matched": status_counts.get("已匹配", 0),
        "unresolved": len(records) - status_counts.get("已匹配", 0),
        "total_amount": round(total_amount, 2),
        "status_counts": dict(status_counts),
    }


def public_record(record: dict[str, Any]) -> dict[str, Any]:
    return {
        key: json_value(value)
        for key, value in record.items()
        if key not in {"original", "date_value"}
    }


def build_match_result(
    ledger_path: str | Path,
    price_path: str | Path,
    requested_priority: list[dict[str, Any]] | None = None,
) -> dict[str, Any]:
    ledger = read_ledger(ledger_path)
    price_book = read_price_book(price_path)
    priority = normalize_priority(price_book["systems"], requested_priority)
    records = match_records(ledger["records"], price_book["items"], priority)
    return {
        "ledger_headers": ledger["headers"],
        "ledger_records": ledger["records"],
        "price_items": price_book["items"],
        "systems": price_book["systems"],
        "priority": priority,
        "records": records,
        "summary": summarize(records),
    }


def reprice_result(result: dict[str, Any], requested_priority: list[dict[str, Any]]) -> dict[str, Any]:
    priority = normalize_priority(result["systems"], requested_priority)
    records = match_records(result["ledger_records"], result["price_items"], priority)
    result["priority"] = priority
    result["records"] = records
    result["summary"] = summarize(records)
    return result


def safe_excel_value(value: Any) -> Any:
    if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
        return "'" + value
    return value


def export_result(
    result: dict[str, Any],
    output_path: str | Path,
    row_ids: list[str] | None = None,
    ledger_name: str = "",
    price_name: str = "",
) -> Path:
    selected_ids = set(row_ids) if row_ids is not None else None
    records = [
        record
        for record in result["records"]
        if selected_ids is None or record["id"] in selected_ids
    ]
    headers = list(result["ledger_headers"])
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "结算汇总"
    detail_sheet = workbook.create_sheet("结算明细")
    summary_sheet.sheet_view.showGridLines = False
    detail_sheet.sheet_view.showGridLines = False

    all_headers = headers + RESULT_HEADERS
    detail_sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(all_headers))
    detail_sheet.cell(1, 1, "项目结算价格匹配结果")
    detail_sheet.cell(1, 1).font = Font(name="Microsoft YaHei", size=16, bold=True, color="FFFFFF")
    detail_sheet.cell(1, 1).fill = PatternFill("solid", fgColor="1F4E78")
    detail_sheet.cell(1, 1).alignment = Alignment(horizontal="center", vertical="center")
    detail_sheet.row_dimensions[1].height = 30

    header_fill = PatternFill("solid", fgColor="DCE6F1")
    thin_border = Border(bottom=Side(style="thin", color="B8C4CE"))
    for column, header in enumerate(all_headers, 1):
        cell = detail_sheet.cell(2, column, header)
        cell.font = Font(name="Microsoft YaHei", size=10, bold=True, color="17324D")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    detail_sheet.row_dimensions[2].height = 28

    result_column = {name: len(headers) + index + 1 for index, name in enumerate(RESULT_HEADERS)}
    status_fills = {
        "已匹配": PatternFill("solid", fgColor="DCFCE7"),
        "未匹配": PatternFill("solid", fgColor="FEE2E2"),
        "字段缺失": PatternFill("solid", fgColor="FEF3C7"),
        "价格表重复": PatternFill("solid", fgColor="FED7AA"),
        "无可用价格": PatternFill("solid", fgColor="EDE9FE"),
    }

    for excel_row, record in enumerate(records, 3):
        original = record["original"]
        for column, header in enumerate(headers, 1):
            value = safe_excel_value(original.get(header))
            cell = detail_sheet.cell(excel_row, column, value)
            if isinstance(value, (date, datetime)):
                cell.number_format = "yyyy-mm-dd"
        detail_sheet.cell(excel_row, result_column["匹配状态"], record["status"])
        detail_sheet.cell(excel_row, result_column["计费项目编号"], safe_excel_value(record["matched_code"]))
        detail_sheet.cell(excel_row, result_column["价格表行"], record["price_row_number"])
        detail_sheet.cell(excel_row, result_column["匹配价格体系"], safe_excel_value(record["selected_system"]))
        detail_sheet.cell(excel_row, result_column["结算单价(元)"], record["selected_price"])
        detail_sheet.cell(excel_row, result_column["结算数量"], record["quantity"])
        price_ref = detail_sheet.cell(excel_row, result_column["结算单价(元)"]).coordinate
        quantity_ref = detail_sheet.cell(excel_row, result_column["结算数量"]).coordinate
        amount_cell = detail_sheet.cell(excel_row, result_column["结算金额(元)"])
        if record["selected_price"] is not None and record["quantity"] is not None:
            amount_cell.value = f'=IF(OR(ISBLANK({price_ref}),ISBLANK({quantity_ref})),"",{price_ref}*{quantity_ref})'
        detail_sheet.cell(excel_row, result_column["匹配状态"]).fill = status_fills.get(record["status"], PatternFill())

    last_row = max(2, len(records) + 2)
    detail_sheet.freeze_panes = "A3"
    detail_sheet.auto_filter.ref = f"A2:{get_column_letter(len(all_headers))}{last_row}"
    detail_sheet.sheet_properties.pageSetUpPr.fitToPage = True
    detail_sheet.page_setup.fitToWidth = 1
    detail_sheet.page_setup.fitToHeight = 0
    for row in detail_sheet.iter_rows(min_row=3, max_row=last_row):
        for cell in row:
            cell.alignment = Alignment(vertical="center")
            cell.border = Border(bottom=Side(style="hair", color="E5E7EB"))
    for header in ("结算单价(元)", "结算数量", "结算金额(元)"):
        detail_sheet.cell(2, result_column[header]).alignment = Alignment(horizontal="right", vertical="center")
        for row in range(3, last_row + 1):
            detail_sheet.cell(row, result_column[header]).number_format = "#,##0.00"

    width_defaults = {
        "委托日期": 12,
        "报告编号": 20,
        "受理编号": 20,
        "报告类别": 18,
        "工程名称": 38,
        "单体工程": 18,
        "委托单位": 28,
        "计费项目": 32,
        "匹配状态": 14,
        "计费项目编号": 16,
        "匹配价格体系": 16,
    }
    for column, header in enumerate(all_headers, 1):
        sample_values = [header]
        for record in records[:200]:
            if header in record["original"]:
                sample_values.append(cell_text(record["original"].get(header)))
            elif header == "匹配状态":
                sample_values.append(record["status"])
            elif header == "计费项目编号":
                sample_values.append(record["matched_code"])
            elif header == "匹配价格体系":
                sample_values.append(record["selected_system"])
        measured = min(max(len(value) for value in sample_values) + 2, 36)
        detail_sheet.column_dimensions[get_column_letter(column)].width = width_defaults.get(header, max(10, measured))

    summary_sheet.merge_cells("A1:F1")
    summary_sheet["A1"] = "项目结算汇总"
    summary_sheet["A1"].font = Font(name="Microsoft YaHei", size=18, bold=True, color="FFFFFF")
    summary_sheet["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    summary_sheet["A1"].alignment = Alignment(horizontal="center")
    summary_sheet.row_dimensions[1].height = 32
    summary_sheet["A3"] = "指标"
    summary_sheet["B3"] = "结果"
    summary_sheet["D3"] = "价格优先级"
    summary_sheet["E3"] = "启用"
    summary_sheet["F3"] = "可用价格数"
    for cell in summary_sheet[3]:
        cell.font = Font(name="Microsoft YaHei", bold=True, color="17324D")
        cell.fill = header_fill

    amount_col_letter = get_column_letter(result_column["结算金额(元)"])
    status_col_letter = get_column_letter(result_column["匹配状态"])
    if records:
        metrics = [
            ("导出记录数", f"=COUNTA('结算明细'!A3:A{last_row})"),
            ("已匹配", f'=COUNTIF(\'结算明细\'!{status_col_letter}3:{status_col_letter}{last_row},"已匹配")'),
            ("待处理", f'=COUNTA(\'结算明细\'!A3:A{last_row})-COUNTIF(\'结算明细\'!{status_col_letter}3:{status_col_letter}{last_row},"已匹配")'),
            ("结算金额合计", f"=SUM('结算明细'!{amount_col_letter}3:{amount_col_letter}{last_row})"),
        ]
    else:
        metrics = [("导出记录数", 0), ("已匹配", 0), ("待处理", 0), ("结算金额合计", 0)]
    for row, (label, value) in enumerate(metrics, 4):
        summary_sheet.cell(row, 1, label)
        summary_sheet.cell(row, 2, value)
    summary_sheet["B7"].number_format = "#,##0.00"

    counts = {system["name"]: system["non_empty"] for system in result["systems"]}
    for row, entry in enumerate(result["priority"], 4):
        summary_sheet.cell(row, 4, f"{row - 3}. {entry['name']}")
        summary_sheet.cell(row, 5, "是" if entry["enabled"] else "否")
        summary_sheet.cell(row, 6, counts.get(entry["name"], 0))

    summary_sheet["A10"] = "台账文件"
    summary_sheet["B10"] = safe_excel_value(ledger_name)
    summary_sheet["A11"] = "价格文件"
    summary_sheet["B11"] = safe_excel_value(price_name)
    summary_sheet["A12"] = "导出时间"
    summary_sheet["B12"] = datetime.now()
    summary_sheet["B12"].number_format = "yyyy-mm-dd hh:mm"
    summary_sheet.column_dimensions["A"].width = 18
    summary_sheet.column_dimensions["B"].width = 38
    summary_sheet.column_dimensions["C"].width = 4
    summary_sheet.column_dimensions["D"].width = 24
    summary_sheet.column_dimensions["E"].width = 10
    summary_sheet.column_dimensions["F"].width = 14
    summary_sheet.freeze_panes = "A3"

    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"
    target = Path(output_path)
    target.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(target)
    return target
