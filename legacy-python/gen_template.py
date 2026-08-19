"""生成外部报价清单导入模板（运行一次即可）。"""
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).resolve().parent.parent
OUTPUT = BASE_DIR / "data" / "报价清单导入模板.xlsx"
OUTPUT.parent.mkdir(exist_ok=True)

# ── 列定义 ────────────────────────────────────────────────
# (表头文字, 宽度, 说明, 必填)
COLUMNS = [
    ("序号",     6,  "同一样品的多行共用同一序号",           False),
    ("样品名称", 20, "材料/构件名称，如：钢筋、预拌砂浆DM", True),
    ("检测项目", 28, "具体检测参数名称，每行填一项",         True),
    ("组/点数",  10, "检测数量（纯数字）",                   False),
    ("单价",     10, "对方报出的单价（纯数字，元）",          False),
    ("合价",     10, "对方报出的合价（纯数字，元）",          False),
    ("备注",     30, "规格、等级或其他说明，如：按三级钢考虑", False),
]

# ── 示例数据（对应图片中的真实格式）────────────────────────
# (序号, 样品名称, 检测项目, 组/点数, 单价, 合价, 备注)
EXAMPLES = [
    (1, "钢筋",       "抗拉、弯曲",       11, 138,  None, "按三级钢考虑"),
    (1, None,         "重量偏差",         None, 59,  None, ""),
    (2, "预拌砂浆DM", "抗压强度",         2,   275, None, ""),
    (2, None,         "凝结时间",         None,180, None, ""),
    (2, None,         "保水率",           None,187, None, ""),
    (2, None,         "稠度损失率",       None,300, None, ""),
    (2, None,         "14d拉伸粘结强度",  None,1800,None, ""),
    (3, "预拌砂浆DP", "抗压强度",         2,   275, None, ""),
    (3, None,         "凝结时间",         None,180, None, ""),
    (3, None,         "保水率",           None,187, None, ""),
]

# ── 颜色常量 ──────────────────────────────────────────────
C_HEADER   = "1E3A5F"   # 深蓝：必填列表头
C_HEADER2  = "2F6496"   # 中蓝：选填列表头
C_SAMPLE   = "FDF3E7"   # 米黄：样品名称行（同图片风格）
C_INST_BG  = "F0F4F8"   # 灰蓝：说明行背景
C_PRICE_HL = "E2EFDA"   # 浅绿：单价列（同图片高亮）
C_BORDER   = "C5D5E8"

thin = Side(style="thin", color=C_BORDER)
border = Border(left=thin, right=thin, top=thin, bottom=thin)


def cell_style(ws, row, col, value, *, bold=False, bg=None, fg="000000",
              align="left", wrap=False, num_fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="微软雅黑", bold=bold, color=fg, size=10)
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    c.border = border
    if num_fmt:
        c.number_format = num_fmt
    return c


# 列索引（1-based）
COL_IDX = {col[0]: i + 1 for i, col in enumerate(COLUMNS)}
PRICE_COL  = COL_IDX["单价"]
TOTAL_COL  = COL_IDX["合价"]
NUM_COLS   = {PRICE_COL, TOTAL_COL, COL_IDX["组/点数"]}


def make_template():
    wb = openpyxl.Workbook()

    # ════ Sheet 1：填写说明 ════════════════════════════════
    ws_info = wb.active
    ws_info.title = "填写说明"
    ws_info.sheet_view.showGridLines = False
    ws_info.column_dimensions["A"].width = 16
    ws_info.column_dimensions["B"].width = 18
    ws_info.column_dimensions["C"].width = 42

    ws_info.merge_cells("A1:C1")
    t = ws_info["A1"]
    t.value = "外部报价清单 · 导入模板说明"
    t.font = Font(name="微软雅黑", bold=True, size=13, color="FFFFFF")
    t.fill = PatternFill("solid", fgColor=C_HEADER)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws_info.row_dimensions[1].height = 32

    for col, h in enumerate(["列名", "是否必填", "填写说明"], 1):
        cell_style(ws_info, 2, col, h, bold=True, bg="2F6496", fg="FFFFFF", align="center")
    ws_info.row_dimensions[2].height = 22

    for r, (name, _, desc, required) in enumerate(COLUMNS, 3):
        req_flag = required
        bg = "FFF8F0" if req_flag else None
        cell_style(ws_info, r, 1, name, bold=req_flag, bg=bg)
        cell_style(ws_info, r, 2, "★ 必填" if req_flag else "选填", bg=bg,
                   fg="C0392B" if req_flag else "555555", align="center")
        cell_style(ws_info, r, 3, desc, bg=bg, wrap=True)
        ws_info.row_dimensions[r].height = 20

    note_row = len(COLUMNS) + 4
    ws_info.merge_cells(f"A{note_row}:C{note_row}")
    cell_style(ws_info, note_row, 1,
               "💡 填写规则：同一样品的多项检测，序号和样品名称只填第一行，后续行留空即可（系统自动补全）。",
               bg="FFFBE6", wrap=True)
    ws_info.row_dimensions[note_row].height = 28

    # ════ Sheet 2：报价清单（模板） ════════════════════════
    ws = wb.create_sheet("报价清单")
    ws.sheet_view.showGridLines = False

    ncols = len(COLUMNS)
    last_col_letter = get_column_letter(ncols)

    # 行1：大标题
    ws.merge_cells(f"A1:{last_col_letter}1")
    title = ws["A1"]
    title.value = "外部报价清单（请在此表填写，勿修改表头）"
    title.font = Font(name="微软雅黑", bold=True, size=12, color="FFFFFF")
    title.fill = PatternFill("solid", fgColor=C_HEADER)
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # 行2：字段提示（灰色小字）
    for col_idx, (_, _, hint, _) in enumerate(COLUMNS, 1):
        cell_style(ws, 2, col_idx, hint, bg=C_INST_BG, fg="667085", wrap=True)
    ws.row_dimensions[2].height = 34

    # 行3：表头
    for col_idx, (name, width, _, required) in enumerate(COLUMNS, 1):
        bg = C_HEADER if required else C_HEADER2
        cell_style(ws, 3, col_idx, name, bold=True, bg=bg, fg="FFFFFF", align="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[3].height = 24

    # 行4起：示例数据
    # 计算每个序号对应的行范围，用于合并单元格
    groups: dict[int, list[int]] = {}  # seq -> [row_indices]
    DATA_START = 4
    for i, row_data in enumerate(EXAMPLES):
        excel_row = DATA_START + i
        seq, sample, param, qty, price, total, remark = row_data
        row_bg = C_SAMPLE  # 统一米黄背景（与图片一致）

        cell_style(ws, excel_row, COL_IDX["序号"],     seq,    bg=row_bg, align="center")
        cell_style(ws, excel_row, COL_IDX["样品名称"], sample or "", bg=row_bg, align="center")
        cell_style(ws, excel_row, COL_IDX["检测项目"], param,  bg=row_bg, align="center")
        cell_style(ws, excel_row, COL_IDX["组/点数"],  qty,    bg=row_bg, align="center", num_fmt="#,##0")
        cell_style(ws, excel_row, COL_IDX["单价"],     price,  bg=C_PRICE_HL, align="center", num_fmt="#,##0.##")
        cell_style(ws, excel_row, COL_IDX["合价"],     total,  bg=row_bg, align="center", num_fmt="#,##0.##")
        cell_style(ws, excel_row, COL_IDX["备注"],     remark, bg=row_bg, wrap=True)
        ws.row_dimensions[excel_row].height = 18

        if seq not in groups:
            groups[seq] = []
        groups[seq].append(excel_row)

    # 合并序号和样品名称列中连续同组的单元格
    for seq, rows in groups.items():
        if len(rows) > 1:
            for merge_col in (COL_IDX["序号"], COL_IDX["样品名称"], COL_IDX["组/点数"]):
                ws.merge_cells(
                    start_row=rows[0], start_column=merge_col,
                    end_row=rows[-1],  end_column=merge_col,
                )
                ws.cell(rows[0], merge_col).alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )

    # 空白填写区（50行）
    blank_start = DATA_START + len(EXAMPLES)
    for r in range(blank_start, blank_start + 50):
        for col_idx, (_, _, _, _) in enumerate(COLUMNS, 1):
            bg = C_PRICE_HL if col_idx == PRICE_COL else None
            fmt = "#,##0.##" if col_idx in NUM_COLS else None
            cell_style(ws, r, col_idx, None, bg=bg, num_fmt=fmt)
        ws.row_dimensions[r].height = 18

    # 冻结前3行
    ws.freeze_panes = "A4"

    wb.save(OUTPUT)
    print(f"模板已生成：{OUTPUT}")


if __name__ == "__main__":
    make_template()
